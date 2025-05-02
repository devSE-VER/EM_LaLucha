import pandas as pd
import re
import shutil
import logging
from pathlib import Path
# from datetime import datetime
# from dotenv import load_dotenv
# from os import getenv
from openpyxl import load_workbook
from typing import List, Dict, Optional
from constants import DIR_GEN, participants
from process.functions import get_files
from process.database import check_data_exist, get_validated_engine


# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler('generation_import.log'),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)


class GenerationFileProcessor:
    PATTERN_FILE = '^LOAD PROFILE LA LUCHA(.*).xlsx'
    DATE_PATTERN = '(\d{2})(\d{2})(\d{4})'
    MONTH_STR = {
        '01': 'Jan', '02': 'Feb', '03': 'Mar', '04': 'Apr',
        '05': 'May', '06': 'Jun', '07': 'Jul', '08': 'Aug',
        '09': 'Sep', '10': 'Oct', '11': 'NOV', '12': 'Dec'
    }

    def __init__(self, engine, dir_gen: str):
        self.engine = engine
        self.dir_gen = Path(dir_gen)
        self.loaded_dir = self.dir_gen / 'loaded'
        self.failed_dir = self.dir_gen / 'failed'
        self._ensure_directories()

    def _ensure_directories(self):
        """Create necessary directories if they don't exist."""
        self.loaded_dir.mkdir(exist_ok=True)
        self.failed_dir.mkdir(exist_ok=True)

    def _parse_date_from_filename(self, filename: str) -> Optional[str]:
        """Extract and parse date from filename."""
        try:
            match = re.search(self.DATE_PATTERN, filename)
            if match:
                day, month, year = match.groups()
                return f'{year}-{month}-{day}'
            return None
        except Exception as e:
            logger.error(f"Error parsing date from filename {filename}: {e}")
            return None

    def _process_excel_file(self, file_path: Path, sheet_name: str) -> Optional[pd.DataFrame]:
        """Process Excel file and return DataFrame with generation data."""
        try:
            wb = load_workbook(file_path, data_only=True)
            ws = wb[sheet_name]

            gen_data = []
            for i in range(3, 27):
                opr_hr = ws[f"I{i}"].value
                gen_mw = round(float(ws[f"K{i}"].value), 2) if ws[f"K{i}"].value else 0.0
                gen_data.append([opr_hr, gen_mw])

            return pd.DataFrame(gen_data, columns=['opr_hr', 'generacion_mw'])
        except Exception as e:
            logger.error(f"Error processing file {file_path}: {e}")
            return None

    def _move_file(self, file_path: Path, success: bool):
        """Move processed file to appropriate directory."""
        target_dir = self.loaded_dir if success else self.failed_dir
        try:
            shutil.move(str(file_path), str(target_dir / file_path.name))
        except Exception as e:
            logger.error(f"Error moving file {file_path}: {e}")

    def process_files(self):
        """Main method to process generation files."""

        process_result = []
        file_list = get_files(self.dir_gen, self.PATTERN_FILE)

        if not file_list:
            logger.info("No files to process")
            return

        for filename in file_list:
            file_path = Path(self.dir_gen) / filename
            logger.info(f"Processing file: {filename}")

            try:
                # Parse date from filename
                opr_dt = self._parse_date_from_filename(filename)
                print(opr_dt)
                if not opr_dt:
                    raise ValueError(f"Could not parse date from filename: {filename}")

                # Check if data already exists
                if check_data_exist(self.engine, schema='dbo', table='Generacion', opr_dt=opr_dt):
                    logger.info(f"Data for {opr_dt} already exists, skipping")
                    # self._move_file(file_path, True)
                    process_result.append([filename, True])
                    continue

                # Get sheet name
                _, month, day = opr_dt.split('-')
                sheet_name = f"{day}-{self.MONTH_STR[month]}"

                # Process file
                df_temp = self._process_excel_file(file_path, sheet_name)
                if df_temp is None:
                    raise ValueError("Error processing Excel file")

                # Add date column
                df_temp['opr_dt'] = opr_dt

                # Reorder columns
                df_gen = df_temp[['opr_dt', 'opr_hr', 'generacion_mw']]

                if df_gen.empty:
                    logger.warning(f"No data found in {filename}")
                    continue

                # Upload to database
                df_gen.to_sql(name='Generacion', con=self.engine,
                              if_exists='append', schema='dbo',
                              index=False, chunksize=1000
                              )

                process_result.append([filename, True])

                logger.info(f"Successfully processed {filename}")

                # self._move_file(file_path, True)

            except Exception as e:
                process_result.append([filename, False])
                logger.error(f"Error processing {filename}: {e}")
                # self._move_file(file_path, False)
        return pd.DataFrame(process_result, columns=['filename', 'result'])


def move_files(file_path: Path, files_df:pd.DataFrame):
    print(files_df)
    """Move processed files to appropriate directory."""
    loaded_dir = Path(file_path) / 'loaded'
    failed_dir = Path(file_path) / 'failed'

    for index, row in files_df.iterrows():
        print(row)
        filename_path = Path(file_path) / row['filename']

        """Move processed file to appropriate directory."""
        target_dir = loaded_dir  if row['all_true'] else failed_dir
        try:
            shutil.move(str(filename_path), str(target_dir / filename_path.name))
        except Exception as e:
            logger.error(f"Error moving file {file_path}: {e}")

def create_pivot_df(df:pd.DataFrame) -> Optional[pd.DataFrame]:
    """Create pivot dataframe with all_true column added."""
    df_pivot = df.pivot_table(
        values='result',  # Values to aggregate
        index=['filename'],  # Row indices
        columns='server',  # Column indices (servers)
        aggfunc='first'  # Aggregation function
    ).reset_index()

    # Add a new column that performs AND operation across all server columns
    server_columns = [col for col in df_pivot.columns if col != 'filename']  # Get all server columns
    df_pivot['all_true'] = df_pivot[server_columns].all(axis=1)  # AND operation

    return df_pivot


def main():
    """Main function to run the generation import process."""
    # Define DataFrame
    df_result = pd.DataFrame(columns=['server', 'filename', 'result'])
    try:
        for pm, pm_data in participants.items():
            logger.info(f"Processing participant: {pm}")

            for srv, srv_data in pm_data['servers'].items():
                if srv_data['ENABLE']:
                    engine = get_validated_engine(srv_data)
                    if engine:
                        logger.info(f"Connected to server: {srv}")
                        processor = GenerationFileProcessor(engine, DIR_GEN)
                        processor_result = processor.process_files()
                        if processor_result is not None:
                            processor_result['server'] = srv
                            df_result = pd.concat([df_result, processor_result], ignore_index=True)
                    else:
                        logger.error(f"Failed to connect to server: {srv}")

        # Move files according df_result
        # Create a pivot table
        logger.info(f"Search for moving files...")
        pivot_df = create_pivot_df(df_result)

        if not pivot_df.empty:
            logger.info(f"Start moving files...")
            move_files(Path(DIR_GEN), pivot_df)

    except Exception as e:
        logger.error(f"Main process error: {e}")
        raise


if __name__ == '__main__':
    main()